import os
import pandas as pd
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import datetime

# Mapeamento para exibir os nomes dos coeficientes formatados no Excel
NOMES_COEFICIENTES = {
    "liquidez_corrente": "Liquidez Corrente",
    "endividamento": "Coeficiente de Endividamento",
    "roe": "ROE (Retorno s/ Patrimônio Líquido)",
    "alavancagem": "Alavancagem Financeira",
}

def _aplicar_formatacao(
    worksheet,
    df: pd.DataFrame,
    formato_moeda: str,
    formato_porcentagem: str,
    start_row: int = 1,
):
    """Aplica formatação numérica de Moeda e Porcentagem nas células do Excel."""
    indicadores_pct = [
        "earning_yield",
        "payout",
        "margem bruta",
        "margem líquida",
        "margem_bruta",
        "margem_liquida",
    ]

    is_index_indicadores = df.index.name in ["Indicador", "indicador"] or (
        isinstance(df.index, pd.Index)
        and any(
            item in ["Lucro Líquido", "Despesa Operacional"] for item in df.index
        )
    )

    if is_index_indicadores:
        for row_idx, idx_val in enumerate(df.index, start=start_row + 1):
            nome_indicador = str(idx_val).strip().lower()

            # Caso 1: Porcentagens
            if (
                nome_indicador in indicadores_pct
                or "margem" in nome_indicador
                or "payout" in nome_indicador
            ):
                for col_idx in range(2, len(df.columns) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 1.0 or cell.value < -1.0:
                            cell.value = cell.value / 100.0
                        cell.number_format = formato_porcentagem

            # Caso 2: Valores Financeiros da DRE (Lucro Líquido, Despesa Operacional, etc.)
            elif (
                "lucro" in nome_indicador
                or "despesa" in nome_indicador
                or "receita" in nome_indicador
                or "custo" in nome_indicador
                or "ebit" in nome_indicador
            ):
                for col_idx in range(2, len(df.columns) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = formato_moeda

    else:
        header_colunas = [df.index.name or "ticker"] + list(df.columns)

        for col_idx, col_name in enumerate(header_colunas, start=1):
            nome_coluna = str(col_name).strip().lower()

            if nome_coluna in ["valor_de_mercado", "divida_liquida", "ebit"]:
                for row_idx in range(start_row + 1, start_row + len(df) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = formato_moeda

            elif (
                nome_coluna in indicadores_pct
                or "margem" in nome_coluna
                or "payout" in nome_coluna
            ):
                for row_idx in range(start_row + 1, start_row + len(df) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 1.0 or cell.value < -1.0:
                            cell.value = cell.value / 100.0
                        cell.number_format = formato_porcentagem


def _ajustar_largura_colunas(worksheet):
    """Ajusta a largura de todas as colunas na aba de acordo com o maior conteúdo."""
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = max(max_len + 4, 15)


def _adicionar_graficos_da_tabela(
    worksheet,
    start_row_dados: int,
    start_row_posicionamento: int,
    df_tabela: pd.DataFrame,
    graficos_por_linha: int = 2,
    graficos_existentes: int = 0,
) -> int:
    """Cria gráficos de linha para cada item/linha de um DataFrame (Indicadores, DRE ou Coeficientes)

    e os organiza em uma grade (grid) abaixo das tabelas.

    Retorna a quantidade total de gráficos já inseridos na aba.
    """
    if df_tabela is None or df_tabela.empty:
        return graficos_existentes

    num_colunas = len(df_tabela.columns)

    # O cabeçalho com as categorias (Anos/Período) fica na linha 'start_row_dados'
    categories = Reference(
        worksheet,
        min_col=2,
        min_row=start_row_dados,
        max_col=num_colunas + 1,
        max_row=start_row_dados,
    )

    # Dimensões e espaçamento do grid no Excel
    LARGURA_COLUNAS_GRAFICO = 10  # 10 colunas de largura por gráfico
    ALTURA_LINHAS_GRAFICO = 18  # ~18 linhas do Excel de altura por gráfico

    for idx, nome_item in enumerate(df_tabela.index):
        # Linha exata do dado no Excel
        row_num = start_row_dados + 1 + idx

        chart = LineChart()
        chart.title = f"Evolução: {nome_item}"
        chart.style = 13
        chart.legend = None

        chart.x_axis.title = "Período"
        chart.x_axis.delete = False

        chart.y_axis.title = "Valor"
        chart.y_axis.delete = False

        # Referência da linha do indicador/item específico
        data = Reference(
            worksheet,
            min_col=1,
            min_row=row_num,
            max_col=num_colunas + 1,
            max_row=row_num,
        )

        chart.add_data(data, titles_from_data=True, from_rows=True)
        chart.set_categories(categories)

        chart.height = 8.5
        chart.width = 16.5

        # Índice acumulado para definir posição em Grid (linhas x colunas)
        idx_global = graficos_existentes + idx
        linha_grid = idx_global // graficos_por_linha
        coluna_grid = idx_global % graficos_por_linha

        # Cálculo da célula do Excel para fixar a quina superior esquerda do gráfico
        linha_excel = (
            start_row_posicionamento + 2 + (linha_grid * ALTURA_LINHAS_GRAFICO)
        )
        coluna_excel_num = 1 + (coluna_grid * LARGURA_COLUNAS_GRAFICO)

        letra_coluna = get_column_letter(coluna_excel_num)
        posicao_grafico = f"{letra_coluna}{linha_excel}"

        worksheet.add_chart(chart, posicao_grafico)

    return graficos_existentes + len(df_tabela)

def extrair_resumo_dre(dre: dict) -> pd.DataFrame | None:
    """Extrai Lucro Líquido e Despesas Operacionais da DRE e formata

    no padrão de colunas do histórico (2021 a 2025 + Atual).
    """
    colunas_anos = ["2021", "2022", "2023", "2024", "2025", "Atual"]

    # Busca direta pelas chaves exatas da DRE
    lucro_liquido = dre.get("LUCRO LÍQUIDO - (R$)", [])
    despesa_op = dre.get("DESPESAS/RECEITAS OPERACIONAIS - (R$)", [])

    if not lucro_liquido or not despesa_op:
        return None

    # Inverte as listas para que o primeiro elemento fique em 2021 e o último em 'Atual'
    ll_ordenado = list(reversed(lucro_liquido[:6]))
    despesa_ordenada = list(reversed(despesa_op[:6]))

    dados = {
        "Lucro Líquido": ll_ordenado,
        "Despesa Operacional": despesa_ordenada,
    }

    df = pd.DataFrame(dados, index=colunas_anos).T
    return df


def salvar_em_excel_por_abas(
    resultados: dict[str, dict], caminho_arquivo: str
):
    if not resultados:
        print("Nenhum dado para salvar no arquivo Excel.")
        return

    diretorio = os.path.dirname(caminho_arquivo)
    if diretorio:
        os.makedirs(diretorio, exist_ok=True)

    dfs_ranking = []

    for ticker, dados in resultados.items():
        df_ey = dados["ey"].copy()
        divida = df_ey["divida_liquida"].fillna(0).values[0]
        vme = df_ey["valor_de_mercado"].values[0]
        ebit = df_ey["ebit"].values[0]

        ey = ebit / (divida + vme)
        df_ey["earning_yield"] = ey
        dfs_ranking.append(df_ey)

    df_ranking = pd.concat(dfs_ranking)
    df_ranking.sort_values(by="earning_yield", ascending=False, inplace=True)

    formato_moeda = 'R$ #,##0.00;[Red]-R$ #,##0.00;"-"'
    formato_porcentagem = "0.00%"
    colunas_desejadas = ["2021", "2022", "2023", "2024", "2025", "Atual"]

    tickers_lista = list(resultados.keys())

    with pd.ExcelWriter(caminho_arquivo, engine="openpyxl") as writer:
        # Ranking
        df_ranking.to_excel(writer, sheet_name="Ranking", index=True)
        ws_ranking = writer.sheets["Ranking"]
        _aplicar_formatacao(
            ws_ranking,
            df_ranking,
            formato_moeda,
            formato_porcentagem,
            start_row=1,
        )
        _ajustar_largura_colunas(ws_ranking)

        # Priorização
        _gerar_aba_priorizacao(writer, tickers_lista)

        # Abas Individuais por Empresa
        for ticker, dados in resultados.items():
            nome_aba = str(ticker)[:31]
            df_ey = dados["ey"]
            df_indicadores = dados.get("indicadores")
            coeficientes = dados.get("coeficientes")
            dre_completa = dados.get("dre")

            # A) Earning Yield
            df_ey.to_excel(writer, sheet_name=nome_aba, startrow=0, index=True)
            ws_empresa = writer.sheets[nome_aba]
            _aplicar_formatacao(
                ws_empresa,
                df_ey,
                formato_moeda,
                formato_porcentagem,
                start_row=1,
            )

            linha_atual = len(df_ey) + 4

            # Marcadores das posições de dados
            start_row_ind = None
            start_row_dre = None
            start_row_coef = None

            df_ind_filtrado = None
            df_resumo_dre = None
            df_coef = None

            # B) Indicadores Fundamentalistas
            if df_indicadores is not None and not df_indicadores.empty:
                df_ind_filtrado = df_indicadores.copy()
                cols_presentes = [
                    c for c in colunas_desejadas if c in df_ind_filtrado.columns
                ]
                if cols_presentes:
                    df_ind_filtrado = df_ind_filtrado[cols_presentes]

                ws_empresa.cell(
                    row=linha_atual - 1,
                    column=1,
                    value="Indicadores Fundamentalistas (2021 - Atual)",
                )

                df_ind_filtrado.to_excel(
                    writer,
                    sheet_name=nome_aba,
                    startrow=linha_atual,
                    index=True,
                )

                _aplicar_formatacao(
                    ws_empresa,
                    df_ind_filtrado,
                    formato_moeda,
                    formato_porcentagem,
                    start_row=linha_atual + 1,
                )

                start_row_ind = linha_atual + 1
                linha_atual += len(df_ind_filtrado) + 4

            # C) Resumo da DRE (Lucro Líquido e Despesas Operacionais)
            if dre_completa:
                df_resumo_dre = extrair_resumo_dre(dre_completa)
                if df_resumo_dre is not None:
                    ws_empresa.cell(
                        row=linha_atual - 1,
                        column=1,
                        value="Resultados de DRE (2021 - Atual)",
                    )

                    df_resumo_dre.to_excel(
                        writer,
                        sheet_name=nome_aba,
                        startrow=linha_atual,
                        index=True,
                    )

                    _aplicar_formatacao(
                        ws_empresa,
                        df_resumo_dre,
                        formato_moeda,
                        formato_porcentagem,
                        start_row=linha_atual + 1,
                    )

                    start_row_dre = linha_atual + 1
                    linha_atual += len(df_resumo_dre) + 4

            # D) Coeficientes e Índices Financeiros
            if coeficientes:
                num_anos = len(next(iter(coeficientes.values())))
                colunas_anos = colunas_desejadas[-num_anos:]

                df_coef = pd.DataFrame(coeficientes, index=colunas_anos).T
                df_coef.rename(index=NOMES_COEFICIENTES, inplace=True)

                ws_empresa.cell(
                    row=linha_atual - 1,
                    column=1,
                    value="Coeficientes e Índices Financeiros",
                )

                df_coef.to_excel(
                    writer,
                    sheet_name=nome_aba,
                    startrow=linha_atual,
                    index=True,
                )

                _aplicar_formatacao(
                    ws_empresa,
                    df_coef,
                    formato_moeda,
                    formato_porcentagem,
                    start_row=linha_atual + 1,
                )

                start_row_coef = linha_atual + 1
                linha_atual += len(df_coef) + 4

            # E) GERAR TODOS OS GRÁFICOS ORGANIZADOS EM GRID
            total_graficos = 0

            # 1. Gráficos de Indicadores
            if df_ind_filtrado is not None and start_row_ind:
                total_graficos = _adicionar_graficos_da_tabela(
                    worksheet=ws_empresa,
                    start_row_dados=start_row_ind,
                    start_row_posicionamento=linha_atual,
                    df_tabela=df_ind_filtrado,
                    graficos_por_linha=2,
                    graficos_existentes=total_graficos,
                )

            # 2. Gráficos de DRE (Lucro Líquido e Despesa Operacional)
            if df_resumo_dre is not None and start_row_dre:
                total_graficos = _adicionar_graficos_da_tabela(
                    worksheet=ws_empresa,
                    start_row_dados=start_row_dre,
                    start_row_posicionamento=linha_atual,
                    df_tabela=df_resumo_dre,
                    graficos_por_linha=2,
                    graficos_existentes=total_graficos,
                )

            # 3. Gráficos de Coeficientes (Liquidez, Endividamento, ROE, Alavancagem)
            if df_coef is not None and start_row_coef:
                total_graficos = _adicionar_graficos_da_tabela(
                    worksheet=ws_empresa,
                    start_row_dados=start_row_coef,
                    start_row_posicionamento=linha_atual,
                    df_tabela=df_coef,
                    graficos_por_linha=2,
                    graficos_existentes=total_graficos,
                )

            _ajustar_largura_colunas(ws_empresa)

    print(
        f"\n[SUCESSO] Arquivo Excel exportado com sucesso para: '{caminho_arquivo}'"
    )


def _gerar_aba_priorizacao(writer, tickers: list[str]):
    wb = writer.book
    ws = wb.create_sheet(title="Priorizacao")

    ws["A1"] = "MATRIZ DE PRIORIZAÇÃO DE COMPRA"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)

    ws["A3"] = "Critério de Pontuação"
    ws["B3"] = "Pontos se Sim"
    ws["C3"] = "Pontos se Não"

    regras = [
        ("1 - Está em recuperação judicial?", -1, 0),
        ("2 - Está no ranking das 20 ações mais baratas?", 1, 0),
        ("3 - Margem bruta constante ou crescente acima de 40%?", 1, 0),
        ("4 - Lucro líquido crescente?", 1, 0),
        ("5 - Lucro por ação crescente?", 1, 0),
        ("6 - Despesa operacional crescente?", -1, 0),
        ("7 - Distribui dividendos?", 1, 0),
        ("8 - Estoque crescente?", 1, 0),
        ("9 - Caixa e equivalentes crescente ou constante?", 1, 0),
        ("10 - Tem mais dívidas de curto prazo que longo prazo?", -1, 0),
        ("11 - Coeficiente de liquidez constante/crescente?", 1, 0),
        ("12 - Coeficiente de endividamento constante/decrescente?", 1, 0),
        ("13 - Lucro acumulado crescente?", 1, 0),
        ("14 - Retorno sobre o PL (ROE) alto/crescente?", 1, 0),
        ("15 - Alavancagem crescente?", -1, 0),
        ("16 - Cotada abaixo do preço teto de Bazin?", 1, 0),
        ("17 - Cotada abaixo do valor intrínseco de Graham?", 1, 0),
    ]

    for idx, (pergunta, p_sim, p_nao) in enumerate(regras, start=4):
        ws.cell(row=idx, column=1, value=pergunta)
        ws.cell(row=idx, column=2, value=p_sim)
        ws.cell(row=idx, column=3, value=p_nao)

    for col in range(1, 4):
        ws.cell(row=3, column=col).font = openpyxl.styles.Font(
            bold=True, color="FFFFFF"
        )
        ws.cell(row=3, column=col).fill = openpyxl.styles.PatternFill(
            start_color="1F497D", end_color="1F497D", fill_type="solid"
        )

    start_row_tabela = 24

    headers_tabela = [
        "Ticker",
        "Q1: Rec. Jud. (-1)",
        "Q2: Top 20 (+1)",
        "Q3: M. Bruta (+1)",
        "Q4: L. Líquido (+1)",
        "Q5: LPA (+1)",
        "Q6: Desp. Cresc. (-1)",
        "Q7: Dividendos (+1)",
        "Q8: Estoque (+1)",
        "Q9: Caixa (+1)",
        "Q10: Dív. Curto Pr. (-1)",
        "Q11: Liquidez (+1)",
        "Q12: Endividamento (+1)",
        "Q13: L. Acumulado (+1)",
        "Q14: ROE (+1)",
        "Q15: Alavancagem (-1)",
        "Q16: Teto Bazin (+1)",
        "Q17: Graham (+1)",
        "Pontuação\nTotal",  # Quebra de linha para economizar largura
    ]

    # Define o índice correto da última coluna (Coluna R / 18)
    col_total_idx = len(headers_tabela)

    for col_idx, header in enumerate(headers_tabela, start=1):
        cell = ws.cell(row=start_row_tabela, column=col_idx, value=header)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    dv = DataValidation(
        type="list", formula1='"Sim,Não"', allow_blank=True, showDropDown=True
    )
    ws.add_data_validation(dv)

    linha_atual = start_row_tabela + 1

    for ticker in tickers:
        ws.cell(row=linha_atual, column=1, value=ticker).font = (
            openpyxl.styles.Font(bold=True)
        )

        partes_formula = []

        # Iterando pelas 17 colunas de perguntas (Colunas B até R / 2 até 18)
        for idx_regra, (pergunta, p_sim, p_nao) in enumerate(regras):
            col_idx = idx_regra + 2
            c_cell = ws.cell(row=linha_atual, column=col_idx, value="Não")
            dv.add(c_cell)

            col_letter = openpyxl.utils.get_column_letter(col_idx)
            termo_formula = (
                f'IF({col_letter}{linha_atual}="Sim",{p_sim},'
                f'IF({col_letter}{linha_atual}="Não",{p_nao},0))'
            )
            partes_formula.append(termo_formula)

        formula_total = "=" + " + ".join(partes_formula)

        # Grava a fórmula na coluna 18 (Coluna R)
        cell_total = ws.cell(
            row=linha_atual, column=col_total_idx, value=formula_total
        )
        cell_total.font = openpyxl.styles.Font(bold=True)
        cell_total.alignment = openpyxl.styles.Alignment(horizontal="center")

        linha_atual += 1

    max_linha = linha_atual - 1
    col_letter_max = openpyxl.utils.get_column_letter(col_total_idx)
    ws.auto_filter.ref = f"A{start_row_tabela}:{col_letter_max}{max_linha}"

    _ajustar_largura_colunas(ws)

    # Trava a largura da coluna de Pontuação Total para 14 para não esticar
    ws.column_dimensions[col_letter_max].width = 14